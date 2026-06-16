"""
ÚLTIMA MODIFICACIÓN: 12/6/2025 por S4NDULOS
PROPÓSITO: Generación de reportes en PDF y Excel
           Soporta listado de productos, stock bajo y movimientos
"""

import io
from datetime import date
from typing import List, Optional
from sqlalchemy.orm import Session
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from fastapi.responses import StreamingResponse

from app.models.producto import ProductoDB
from app.models.movimiento import MovimientoDB
from app.services.producto_service import get_all_productos, get_productos_stock_bajo
from app.services.movimiento_service import get_movimientos

# ------------------------------------------------------------
# UTILIDADES COMUNES
# ------------------------------------------------------------

def _generar_excel_generico(titulo: str, encabezados: List[str], datos: List[List]) -> io.BytesIO:
    """Genera un archivo Excel en memoria a partir de datos tabulares."""
    wb = Workbook()
    ws = wb.active
    ws.title = titulo[:31]  # Excel limita a 31 caracteres

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")

    # Escribir encabezados
    for col, header in enumerate(encabezados, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment

    # Escribir datos
    for row_idx, fila in enumerate(datos, 2):
        for col_idx, valor in enumerate(fila, 1):
            ws.cell(row=row_idx, column=col_idx, value=valor)

    # Ajustar anchos de columna
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[col_letter].width = adjusted_width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _generar_pdf_generico(titulo: str, encabezados: List[str], datos: List[List]) -> io.BytesIO:
    """Genera un archivo PDF en memoria a partir de datos tabulares."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
    elements = []

    # Estilo título
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'TitleStyle',
        parent=styles['Heading1'],
        fontSize=16,
        alignment=1,  # centrado
        spaceAfter=20
    )
    elements.append(Paragraph(titulo, title_style))
    elements.append(Spacer(1, 0.2 * inch))

    # Construir tabla
    tabla_data = [encabezados] + datos
    tabla = Table(tabla_data, repeatRows=1)
    tabla.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(tabla)

    doc.build(elements)
    buffer.seek(0)
    return buffer


# ------------------------------------------------------------
# REPORTE DE PRODUCTOS
# ------------------------------------------------------------

def generar_reporte_productos(db: Session, formato: str) -> StreamingResponse:
    productos = get_all_productos(db, skip=0, limit=1000)  # sin paginación para reporte
    encabezados = ["ID", "Nombre", "Precio", "Stock", "Stock Mínimo", "Stock Máximo"]
    datos = [
        [
            p.id,
            p.nombre,
            f"${p.precio:.2f}",
            p.stock,
            p.stock_minimo,
            p.stock_maximo
        ]
        for p in productos
    ]

    titulo = "Reporte de Productos"

    if formato == "excel":
        buffer = _generar_excel_generico(titulo, encabezados, datos)
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        filename = "productos.xlsx"
    else:
        buffer = _generar_pdf_generico(titulo, encabezados, datos)
        media_type = "application/pdf"
        filename = "productos.pdf"

    return StreamingResponse(
        buffer,
        media_type=media_type,
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ------------------------------------------------------------
# REPORTE DE STOCK BAJO
# ------------------------------------------------------------

def generar_reporte_stock_bajo(db: Session, formato: str, umbral: Optional[int] = None) -> StreamingResponse:
    productos = get_productos_stock_bajo(db, umbral)
    encabezados = ["ID", "Nombre", "Stock Actual", "Stock Mínimo", "Stock Máximo", "Precio"]
    datos = [
        [p.id, p.nombre, p.stock, p.stock_minimo, p.stock_maximo, f"${p.precio:.2f}"]
        for p in productos
    ]

    titulo = f"Reporte de Stock Bajo (umbral: {umbral if umbral else 'por producto'})"

    if formato == "excel":
        buffer = _generar_excel_generico(titulo, encabezados, datos)
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        filename = "stock_bajo.xlsx"
    else:
        buffer = _generar_pdf_generico(titulo, encabezados, datos)
        media_type = "application/pdf"
        filename = "stock_bajo.pdf"

    return StreamingResponse(
        buffer,
        media_type=media_type,
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ------------------------------------------------------------
# REPORTE DE MOVIMIENTOS
# ------------------------------------------------------------

def generar_reporte_movimientos(
    db: Session,
    formato: str,
    fecha_desde: Optional[date] = None,
    fecha_hasta: Optional[date] = None,
    producto_id: Optional[int] = None
) -> StreamingResponse:
    # Obtener movimientos (sin paginación, máximo 5000 para evitar saturación)
    movimientos = get_movimientos(db, skip=0, limit=5000, producto_id=producto_id, tipo=None)
    
    # Filtrar por fechas manualmente porque get_movimientos no soporta fechas
    if fecha_desde or fecha_hasta:
        filtrados = []
        for m in movimientos:
            m_date = m.fecha_hora.date()
            if fecha_desde and m_date < fecha_desde:
                continue
            if fecha_hasta and m_date > fecha_hasta:
                continue
            filtrados.append(m)
        movimientos = filtrados

    encabezados = ["ID Mov.", "Producto ID", "Producto Nombre", "Tipo", "Cantidad", "Stock Resultante", "Usuario ID", "Fecha/Hora"]
    datos = []
    for m in movimientos:
        # Obtener nombre del producto (puede ser None si fue eliminado)
        nombre_producto = db.query(ProductoDB.nombre).filter(ProductoDB.id == m.producto_id).scalar()
        datos.append([
            m.id,
            m.producto_id,
            nombre_producto if nombre_producto else "Eliminado",
            m.tipo,
            m.cantidad,
            m.stock_resultante,
            m.usuario_id if m.usuario_id else "Sistema",
            m.fecha_hora.strftime("%Y-%m-%d %H:%M:%S")
        ])

    titulo = "Reporte de Movimientos"
    if producto_id:
        titulo += f" - Producto ID {producto_id}"
    if fecha_desde:
        titulo += f" desde {fecha_desde}"
    if fecha_hasta:
        titulo += f" hasta {fecha_hasta}"

    if formato == "excel":
        buffer = _generar_excel_generico(titulo, encabezados, datos)
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        filename = "movimientos.xlsx"
    else:
        buffer = _generar_pdf_generico(titulo, encabezados, datos)
        media_type = "application/pdf"
        filename = "movimientos.pdf"

    return StreamingResponse(
        buffer,
        media_type=media_type,
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )